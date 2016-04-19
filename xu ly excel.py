from openpyxl import * # cai dat thu vien openpyxl bang cach:  pip install openpyxl

#wb = load_workbook('demo data.xlsx')
wb = Workbook() # tao mot workbook moi dat ten la ws (object)
ws = wb.active # su dung worksheet dau tien nhin thay (default worksheet) dat ten la ws (la mot object)
ws['A6'] = 42
wb.save('sample_book.xlsx') 
